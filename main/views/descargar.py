import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from ..models import Dependencias  # Importa tu modelo Dependencias aquí
import json

def descargar_excel(request):
    datos_por_secretaria = request.GET.get('datos_por_secretaria')
    print(datos_por_secretaria)
    if not datos_por_secretaria:
        return HttpResponse("No se recibieron datos", status=400)
 
    try:
        datos = eval(datos_por_secretaria)
    except:
        
        print(datos_por_secretaria[0])
    #print(datos[0])
    # Crear el libro y la hoja de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Escribir los encabezados
    headers = [
        "Dependencia", "Almacenamientos", "Conmutadores", "Energías", "Enlaces Dependientes",
        "Enlaces no Dependientes", "Equipos Personales", "Equipos Servidores", "Equipos Télefonicos",
        "Herramientas de Desarrollo", "Sistema de información", "Sistema de información móvil",
        "Impresoras", "Proyectores", "Enlaces", "Site", "Drones", "Firewalls", "Routers"
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    # Escribir los datos
    for row_num, dependencia in enumerate(datos, 2):
        ws[f'A{row_num}'] = dependencia.get('nombre_secretaria', dependencia.get('nombre_dependencia', ''))
        ws[f'B{row_num}'] = dependencia.get('num_almacenamientos', 0)
        ws[f'C{row_num}'] = dependencia.get('num_conmutadores', 0)
        ws[f'D{row_num}'] = dependencia.get('num_energias', 0)
        ws[f'E{row_num}'] = dependencia.get('num_enlaces_dep', 0)
        ws[f'F{row_num}'] = dependencia.get('num_enlaces_nodep', 0)
        ws[f'G{row_num}'] = dependencia.get('num_equipos_personales', 0)
        ws[f'H{row_num}'] = dependencia.get('num_equipos_servidores', 0)
        ws[f'I{row_num}'] = dependencia.get('num_equipos_telefonico', 0)
        ws[f'J{row_num}'] = dependencia.get('num_herramientas_desa', 0)
        ws[f'K{row_num}'] = dependencia.get('num_sistemas_info', 0)
        ws[f'L{row_num}'] = dependencia.get('num_sistemas_info_movil', 0)
        ws[f'M{row_num}'] = dependencia.get('num_impresoras', 0)
        ws[f'N{row_num}'] = dependencia.get('num_proyectores', 0)
        ws[f'O{row_num}'] = dependencia.get('num_enlaces', 0)
        ws[f'P{row_num}'] = dependencia.get('num_sitios', 0)
        ws[f'Q{row_num}'] = dependencia.get('num_drones', 0)
        ws[f'R{row_num}'] = dependencia.get('num_firewalls', 0)
        ws[f'S{row_num}'] = dependencia.get('num_routers', 0)

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte.xlsx'
    wb.save(response)

    return response
