from django.shortcuts import render
from django.db.models import Count, Case, When, Subquery, OuterRef, Q,Prefetch
from ..models import *
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl import Workbook

@login_required
def principal(request):
    selected_secretaria = request.GET.get('secretaria')
    selected_dependencia = request.GET.get('dependencia')
    selected_year = request.GET.get('year')

    nombre_secretaria="Todas"
    if selected_secretaria and selected_secretaria != "None":
        nombre_secretaria = Secretarias.objects.filter(id=selected_secretaria).values_list('nombre_secretaria', flat=True).first()

    #print(selected_secretaria+" "+nombre_secretaria)
    secretarias = Secretarias.objects.all()
    datos_por_secretaria = None

    if selected_secretaria and selected_secretaria != "None":
        if selected_year:
            datos_por_secretaria = Dependencias.objects.filter(id_secretaria=selected_secretaria).annotate(
                num_enlaces_dep = Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    enlaces__fecha__year=selected_year,
                    enlaces__id_tipo=1,
                    enlaces__isnull=False
                ).values('id').annotate(count=Count('enlaces')).values('count')[:1]
                ),
                num_enlaces_nodep = Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    enlaces__fecha__year=selected_year,
                    enlaces__id_tipo=2,
                    enlaces__isnull=False
                ).values('id').annotate(count=Count('enlaces')).values('count')[:1]
                ),
                num_sistemas_info_movil = Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    sistemadeinformacionmovil__fecha__year=selected_year,
                    sistemadeinformacionmovil__isnull=False
                ).values('id').annotate(count=Count('sistemadeinformacionmovil')).values('count')[:1]
                ),
                num_sistemas_info = Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        sistemasinformacion__fecha__year=selected_year,
                        sistemasinformacion__isnull=False
                    ).values('id').annotate(count=Count('sistemasinformacion')).values('count')[:1]
                ),
                num_herramientas_desa = Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    herramientadedesarrollo__fecha__year=selected_year,
                    herramientadedesarrollo__isnull=False
                ).values('id').annotate(count=Count('herramientadedesarrollo')).values('count')[:1]
                ),
                num_equipos_telefonico=Subquery(
                Dependencias.objects.filter(
                        id=OuterRef('id'),
                        equipotelefonico__fecha__year=selected_year,
                        equipotelefonico__isnull=False
                    ).values('id').annotate(count=Count('equipotelefonico')).values('count')[:1]
                ),
                num_almacenamientos=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        almacenamientos__fecha__year=selected_year,
                        almacenamientos__isnull=False
                    ).values('id').annotate(count=Count('almacenamientos')).values('count')[:1]
                ),
                num_conmutadores=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        conmutadores__fecha__year=selected_year,
                        conmutadores__isnull=False
                    ).values('id').annotate(count=Count('conmutadores')).values('count')[:1]
                ),
                num_energias=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        energias__fecha__year=selected_year,
                        energias__isnull=False
                    ).values('id').annotate(count=Count('energias')).values('count')[:1]
                ),
                num_equipos_personales=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        equipospersonales__fecha__year=selected_year,
                        equipospersonales__isnull=False
                    ).values('id').annotate(count=Count('equipospersonales')).values('count')[:1]
                ),
                num_equipos_servidores=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        equiposservidores__fecha__year=selected_year,
                        equiposservidores__isnull=False
                    ).values('id').annotate(count=Count('equiposservidores')).values('count')[:1]
                ),
                num_impresoras=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        impresoras__fecha__year=selected_year,
                        impresoras__isnull=False
                    ).values('id').annotate(count=Count('impresoras')).values('count')[:1]
                ),
                num_proyectores=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        proyectores__fecha__year=selected_year,
                        proyectores__isnull=False
                    ).values('id').annotate(count=Count('proyectores')).values('count')[:1]
                ),
                num_enlaces=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        enlaces__fecha__year=selected_year,
                        enlaces__isnull=False
                    ).values('id').annotate(count=Count('enlaces')).values('count')[:1]
                ),
                num_drones=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        drones__fecha__year=selected_year,
                        drones__isnull=False
                    ).values('id').annotate(count=Count('drones')).values('count')[:1]
                ),
                num_firewalls=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        firewalls__fecha__year=selected_year,
                        firewalls__isnull=False
                    ).values('id').annotate(count=Count('firewalls')).values('count')[:1]
                ),
                num_routers=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        routers__fecha__year=selected_year,
                        routers__isnull=False
                    ).values('id').annotate(count=Count('routers')).values('count')[:1]
                ),
            )
        
        else:
            #print(Dependencias.objects.filter(id_secretaria=selected_secretaria))
            datos_por_secretaria = Dependencias.objects.filter(id_secretaria=selected_secretaria).annotate(
            num_enlaces_dep = Subquery(
            Dependencias.objects.filter(
                id=OuterRef('id'),
                enlaces__id_tipo=1,
                enlaces__isnull=False
            ).values('id').annotate(count=Count('enlaces')).values('count')[:1]
            ),
            num_enlaces_nodep = Subquery(
            Dependencias.objects.filter(
                id=OuterRef('id'),
                enlaces__id_tipo=2,
                enlaces__isnull=False
            ).values('id').annotate(count=Count('enlaces')).values('count')[:1]
            ),
            num_sistemas_info_movil = Subquery(
            Dependencias.objects.filter(
                id=OuterRef('id'),
                sistemadeinformacionmovil__isnull=False
            ).values('id').annotate(count=Count('sistemadeinformacionmovil')).values('count')[:1]
            ),
            num_sistemas_info = Subquery(
            Dependencias.objects.filter(
                id=OuterRef('id'),
                sistemasinformacion__isnull=False
            ).values('id').annotate(count=Count('sistemasinformacion')).values('count')[:1]
            ),
            num_herramientas_desa = Subquery(
            Dependencias.objects.filter(
                id=OuterRef('id'),
                herramientadedesarrollo__isnull=False
            ).values('id').annotate(count=Count('herramientadedesarrollo')).values('count')[:1]
            ),
            num_equipos_telefonico=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    equipotelefonico__isnull=False
                ).values('id').annotate(count=Count('equipotelefonico')).values('count')[:1]
            ),
            num_almacenamientos=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    almacenamientos__isnull=False
                ).values('id').annotate(count=Count('almacenamientos')).values('count')[:1]
            ),
            num_conmutadores=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    conmutadores__isnull=False
                ).values('id').annotate(count=Count('conmutadores')).values('count')[:1]
            ),
            num_energias=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    energias__isnull=False
                ).values('id').annotate(count=Count('energias')).values('count')[:1]
            ),
            num_equipos_personales=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    equipospersonales__isnull=False
                ).values('id').annotate(count=Count('equipospersonales')).values('count')[:1]
            ),
            num_equipos_servidores=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    equiposservidores__isnull=False
                ).values('id').annotate(count=Count('equiposservidores')).values('count')[:1]
            ),
            num_impresoras=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    impresoras__isnull=False
                ).values('id').annotate(count=Count('impresoras')).values('count')[:1]
            ),
            num_proyectores=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    proyectores__isnull=False
                ).values('id').annotate(count=Count('proyectores')).values('count')[:1]
            ),
            num_enlaces=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    enlaces__isnull=False
                ).values('id').annotate(count=Count('enlaces')).values('count')[:1]
            ),
            num_drones=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    drones__isnull=False
                ).values('id').annotate(count=Count('drones')).values('count')[:1]
            ),
            num_firewalls=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    firewalls__isnull=False
                ).values('id').annotate(count=Count('firewalls')).values('count')[:1]
            ),
            num_routers=Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    routers__isnull=False
                ).values('id').annotate(count=Count('routers')).values('count')[:1]
            ),
        )
    elif selected_year:
            datos_por_secretaria = Secretarias.objects.annotate(
                num_enlaces_dep = Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    enlaces__fecha__year=selected_year,
                    enlaces__id_tipo=1,
                    enlaces__isnull=False
                ).values('id').annotate(count=Count('enlaces')).values('count')[:1]
                ),
                num_enlaces_nodep = Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    enlaces__fecha__year=selected_year,
                    enlaces__id_tipo=2,
                    enlaces__isnull=False
                ).values('id').annotate(count=Count('enlaces')).values('count')[:1]
                ),
                num_sistemas_info_movil = Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    sistemadeinformacionmovil__fecha__year=selected_year,
                    sistemadeinformacionmovil__isnull=False
                ).values('id').annotate(count=Count('sistemadeinformacionmovil')).values('count')[:1]
                ),
                num_sistemas_info = Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    sistemasinformacion__fecha__year=selected_year,
                    sistemasinformacion__isnull=False
                ).values('id').annotate(count=Count('sistemasinformacion')).values('count')[:1]
                ),
                num_herramientas_desa = Subquery(
                Dependencias.objects.filter(
                    id=OuterRef('id'),
                    herramientadedesarrollo__fecha__year=selected_year,
                    herramientadedesarrollo__isnull=False
                ).values('id').annotate(count=Count('herramientadedesarrollo')).values('count')[:1]
                ),
                num_equipos_telefonico=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        equipotelefonico__fecha__year=selected_year,
                        equipotelefonico__isnull=False
                    ).values('id').annotate(count=Count('equipotelefonico')).values('count')[:1]
                ),
                num_almacenamientos=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        almacenamientos__fecha__year=selected_year,
                        almacenamientos__isnull=False
                    ).values('id').annotate(count=Count('almacenamientos')).values('count')[:1]
                ),
                num_conmutadores=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        conmutadores__fecha__year=selected_year,
                        conmutadores__isnull=False
                    ).values('id').annotate(count=Count('conmutadores')).values('count')[:1]
                ),
                num_energias=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        energias__fecha__year=selected_year,
                        energias__isnull=False
                    ).values('id').annotate(count=Count('energias')).values('count')[:1]
                ),
                num_equipos_personales=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        equipospersonales__fecha__year=selected_year,
                        equipospersonales__isnull=False
                    ).values('id').annotate(count=Count('equipospersonales')).values('count')[:1]
                ),
                num_equipos_servidores=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        equiposservidores__fecha__year=selected_year,
                        equiposservidores__isnull=False
                    ).values('id').annotate(count=Count('equiposservidores')).values('count')[:1]
                ),
                num_impresoras=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        impresoras__fecha__year=selected_year,
                        impresoras__isnull=False
                    ).values('id').annotate(count=Count('impresoras')).values('count')[:1]
                ),
                num_proyectores=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        proyectores__fecha__year=selected_year,
                        proyectores__isnull=False
                    ).values('id').annotate(count=Count('proyectores')).values('count')[:1]
                ),
                num_enlaces=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        enlaces__fecha__year=selected_year,
                        enlaces__isnull=False
                    ).values('id').annotate(count=Count('enlaces')).values('count')[:1]
                ),
                num_drones=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        drones__fecha__year=selected_year,
                        drones__isnull=False
                    ).values('id').annotate(count=Count('drones')).values('count')[:1]
                ),
                num_firewalls=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        firewalls__fecha__year=selected_year,
                        firewalls__isnull=False
                    ).values('id').annotate(count=Count('firewalls')).values('count')[:1]
                ),
                num_routers=Subquery(
                    Dependencias.objects.filter(
                        id=OuterRef('id'),
                        routers__fecha__year=selected_year,
                        routers__isnull=False
                    ).values('id').annotate(count=Count('routers')).values('count')[:1]
                ),
            )
    else:
        datos_por_secretaria = []  # Inicializa datos_por_secretaria como una lista vacía
        datos = Secretarias.objects.values('nombre_secretaria').annotate(
        num_conmutadores=Count('dependencias__conmutadores', distinct=True),
    )
        for dato in datos:
            datos_por_secretaria.append({
                'nombre_secretaria': dato['nombre_secretaria'],
                'num_conmutadores': dato['num_conmutadores'],
            })

        datos_enlaces_dep = Secretarias.objects.values('nombre_secretaria').annotate(num_enlaces_dep=Count('dependencias__enlaces', filter=Q(dependencias__enlaces__id_tipo=1), distinct=True),)
        for i, dato in enumerate(datos_enlaces_dep):    datos_por_secretaria[i]['num_enlaces_dep'] = dato['num_enlaces_dep']

        datos_enlaces_nodep = Secretarias.objects.values('nombre_secretaria').annotate(num_enlaces_nodep=Count('dependencias__enlaces', filter=Q(dependencias__enlaces__id_tipo=2), distinct=True),)
        for i, dato in enumerate(datos_enlaces_nodep):    datos_por_secretaria[i]['num_enlaces_nodep'] = dato['num_enlaces_nodep']

        datos_sistemas_info_movil = Secretarias.objects.values('nombre_secretaria').annotate(num_sistemas_info_movil=Count('dependencias__sistemadeinformacionmovil', distinct=True),)
        for i, dato in enumerate(datos_sistemas_info_movil):    datos_por_secretaria[i]['num_sistemas_info_movil'] = dato['num_sistemas_info_movil']
        
        datos_sistemas_info = Secretarias.objects.values('nombre_secretaria').annotate(num_sistemas_info=Count('dependencias__sistemasinformacion', distinct=True),)
        for i, dato in enumerate(datos_sistemas_info):  datos_por_secretaria[i]['num_sistemas_info'] = dato['num_sistemas_info']

        datos_herramientas_desa = Secretarias.objects.values('nombre_secretaria').annotate(num_herramientas_desa=Count('dependencias__herramientadedesarrollo', distinct=True),)
        for i, dato in enumerate(datos_herramientas_desa):  datos_por_secretaria[i]['num_herramientas_desa'] = dato['num_herramientas_desa']

        datos_almacenamientos = Secretarias.objects.values('nombre_secretaria').annotate(num_almacenamientos=Count('dependencias__almacenamientos', distinct=True),)
        for i, dato in enumerate(datos_almacenamientos):   datos_por_secretaria[i]['num_almacenamientos'] = dato['num_almacenamientos']

        datos_equipos_telefonico = Secretarias.objects.values('nombre_secretaria').annotate(num_equipos_telefonico=Count('dependencias__equipotelefonico', distinct=True),)
        for i, dato in enumerate(datos_equipos_telefonico):   datos_por_secretaria[i]['num_equipos_telefonico'] = dato['num_equipos_telefonico']

        datos_energias = Secretarias.objects.values('nombre_secretaria').annotate(num_energias=Count('dependencias__energias', distinct=True),)
        for i, dato in enumerate(datos_energias):   datos_por_secretaria[i]['num_energias'] = dato['num_energias']

        datos_equipos_personales = Secretarias.objects.values('nombre_secretaria').annotate(num_equipos_personales=Count('dependencias__equipospersonales', distinct=True),)
        for i, dato in enumerate(datos_equipos_personales):   datos_por_secretaria[i]['num_equipos_personales'] = dato['num_equipos_personales']

        datos_equipos_servidores = Secretarias.objects.values('nombre_secretaria').annotate(num_equipos_servidores=Count('dependencias__equiposservidores', distinct=True),)
        for i, dato in enumerate(datos_equipos_servidores):   datos_por_secretaria[i]['num_equipos_servidores'] = dato['num_equipos_servidores']

        datos_impresoras = Secretarias.objects.values('nombre_secretaria').annotate(num_impresoras=Count('dependencias__impresoras', distinct=True),)
        for i, dato in enumerate(datos_impresoras):   datos_por_secretaria[i]['num_impresoras'] = dato['num_impresoras']

        datos_proyectores = Secretarias.objects.values('nombre_secretaria').annotate(num_proyectores=Count('dependencias__proyectores', distinct=True),)
        for i, dato in enumerate(datos_proyectores):   datos_por_secretaria[i]['num_proyectores'] = dato['num_proyectores']

        datos_enlaces = Secretarias.objects.values('nombre_secretaria').annotate(num_enlaces=Count('dependencias__enlaces', distinct=True),)
        for i, dato in enumerate(datos_enlaces):   datos_por_secretaria[i]['num_enlaces'] = dato['num_enlaces']

        datos_drones = Secretarias.objects.values('nombre_secretaria').annotate(num_drones=Count('dependencias__drones', distinct=True),)
        for i, dato in enumerate(datos_drones):   datos_por_secretaria[i]['num_drones'] = dato['num_drones']

        datos_firewalls = Secretarias.objects.values('nombre_secretaria').annotate(num_firewalls=Count('dependencias__firewalls', distinct=True),)
        for i, dato in enumerate(datos_firewalls):   datos_por_secretaria[i]['num_firewalls'] = dato['num_firewalls']

        datos_routers = Secretarias.objects.values('nombre_secretaria').annotate(num_routers=Count('dependencias__routers', distinct=True),)
        for i, dato in enumerate(datos_routers):   datos_por_secretaria[i]['num_routers'] = dato['num_routers']

    lista_diccionarios = []

    if not (isinstance(datos_por_secretaria, dict) or isinstance(datos_por_secretaria, list) or isinstance(datos_por_secretaria, tuple)):
    # Itera sobre la lista de objetos Dependencias y conviértelos a diccionarios
        for dependencia in datos_por_secretaria:
            # Convierte el objeto Dependencias a un diccionario
            dependencia_dict = dependencia.__dict__
            dependencia_dict.pop('_state', None)

            lista_diccionarios.append(dependencia_dict)

    else:
        lista_diccionarios = datos_por_secretaria
    
    print(lista_diccionarios)
    #print(datos_por_secretaria)             
    if 'export_excel' in request.GET:
        datos_por_secretaria = list(lista_diccionarios)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=datos.xlsx'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"

        columns = [
            'Nombre', 'Num Conmutadores', 'Num Enlaces Dep',
            'Num Enlaces No Dep', 'Num Sistemas Info Movil', 'Num Sistemas Info', 
            'Num Herramientas Desa', 'Num Almacenamientos', 'Num Equipos Telefonico', 
            'Num Energias', 'Num Equipos Personales', 'Num Equipos Servidores', 
            'Num Impresoras', 'Num Proyectores', 'Num Enlaces', 'Num Drones', 
            'Num Firewalls', 'Num Routers'
        ]
        
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=row_num, column=col_num, value=column_title)

        for dato in datos_por_secretaria:
            row_num += 1
            nombre = dato.get('nombre_secretaria') or dato.get('nombre_dependencia')
            ws.cell(row=row_num, column=1, value=nombre)
            ws.cell(row=row_num, column=2, value=int(dato.get('num_conmutadores') or 0))
            ws.cell(row=row_num, column=3, value=int(dato.get('num_enlaces_dep') or 0))
            ws.cell(row=row_num, column=4, value=int(dato.get('num_enlaces_nodep') or 0))
            ws.cell(row=row_num, column=5, value=int(dato.get('num_sistemas_info_movil') or 0))
            ws.cell(row=row_num, column=6, value=int(dato.get('num_sistemas_info') or 0))
            ws.cell(row=row_num, column=7, value=int(dato.get('num_herramientas_desa') or 0))
            ws.cell(row=row_num, column=8, value=int(dato.get('num_almacenamientos') or 0))
            ws.cell(row=row_num, column=9, value=int(dato.get('num_equipos_telefonico') or 0))
            ws.cell(row=row_num, column=10, value=int(dato.get('num_energias') or 0))
            ws.cell(row=row_num, column=11, value=int(dato.get('num_equipos_personales') or 0))
            ws.cell(row=row_num, column=12, value=int(dato.get('num_equipos_servidores') or 0))
            ws.cell(row=row_num, column=13, value=int(dato.get('num_impresoras') or 0))
            ws.cell(row=row_num, column=14, value=int(dato.get('num_proyectores') or 0))
            ws.cell(row=row_num, column=15, value=int(dato.get('num_enlaces') or 0))
            ws.cell(row=row_num, column=16, value=int(dato.get('num_drones') or 0))
            ws.cell(row=row_num, column=17, value=int(dato.get('num_firewalls') or 0))
            ws.cell(row=row_num, column=18, value=int(dato.get('num_routers') or 0))

        wb.save(response)
        return response

    return render(request, 'principal.html', {
        'secretarias': secretarias,
        'datos_por_secretaria': datos_por_secretaria,
        'selected_secretaria': selected_secretaria,
        'selected_dependencia': selected_dependencia,
        'nombre_secretaria':nombre_secretaria,
        'selected_year': selected_year,
    })
